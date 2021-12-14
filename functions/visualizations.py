import json
from io import BytesIO
from typing import Optional

# clustergrammer2 via sklearn uses np.float which is deprecated as of numpy==1.20
import warnings

warnings.filterwarnings(
    action="ignore", category=DeprecationWarning, module="scikit-learn"
)
import pandas as pd

from clustergrammer2 import Network as CGNetwork
from deepdiff import DeepSearch
from openpyxl import load_workbook
from cidc_api.models import DownloadableFiles, prism, TrialMetadata

from .util import (
    BackgroundContext,
    extract_pubsub_data,
    sqlalchemy_session,
    get_blob_as_stream,
)


# sets the maximum number of divisions within a category
## that is shown on the top of the clustergrammer
CLUSTERGRAMMER_MAX_CATEGORY_CARDINALITY = 5


def vis_preprocessing(event: dict, context: BackgroundContext):
    with sqlalchemy_session() as session:
        object_url = extract_pubsub_data(event)
        file_record: DownloadableFiles = DownloadableFiles.get_by_object_url(
            object_url, session=session
        )

        if not file_record:
            raise Exception(f"No downloadable file with object URL {object_url} found.")

        metadata_df = _get_metadata_df(file_record.trial_id)

        # Apply the transformations and get derivative data for visualization.
        for transform_name, transform in _get_transforms().items():
            vis_json = transform(file_record, metadata_df)
            if vis_json:
                # Add the vis config to the file_record
                setattr(file_record, transform_name, vis_json)

        # Save the derivative data additions to the database.
        session.commit()


def _get_metadata_df(trial_id: str) -> pd.DataFrame:
    """
    Build a dataframe containing the participant/sample metadata for this trial,
    joined on CIMAC ID and indexed on CIMAC ID.
    """
    participants_blob = get_blob_as_stream(
        f"{trial_id}/participants.csv", as_string=True
    )
    samples_blob = get_blob_as_stream(f"{trial_id}/samples.csv", as_string=True)

    participants_df = pd.read_csv(participants_blob)
    samples_df = pd.read_csv(samples_blob)

    metadata_df = pd.merge(
        participants_df,
        samples_df,
        left_on="cimac_participant_id",
        right_on="participants.cimac_participant_id",
        how="outer",
    )
    metadata_df.set_index("cimac_id", inplace=True)

    return metadata_df


def _get_transforms() -> dict:
    """ 
    Get a list of functions taking an open file and
    that file's downloadable file record as arguments, returning
    a JSON blob that the frontend will use for visualization.
    """
    return {
        "clustergrammer": _ClustergrammerTransform(),
        "ihc_combined_plot": _ihc_combined_transform,
        "additional_metadata": _add_antibody_metadata,
    }


def _add_antibody_metadata(
    file_record: DownloadableFiles, metadata_df: pd.DataFrame
) -> Optional[dict]:
    """
    Pseudo transformation to add antibody data to the DownloadableFiles.additional_metadata JSON
    Only for upload_type in [cytof, elisa, ihc, micsss, and mif]
    """
    transforms = {
        "cytof": _cytof_antibody_md,
        "elisa": _elisa_antibody_md,
        "ihc": _ihc_antibody_md,
        "micsss": _micsss_antibody_md,
        "mif": _mif_antibody_md,
    }
    upload_type = file_record.upload_type.lower()
    if upload_type not in transforms.keys():
        return None

    with sqlalchemy_session() as session:
        ct_md = TrialMetadata.find_by_trial_id(
            file_record.trial_id, session=session
        ).metadata_json

    assay_instances = ct_md.get("assays", {}).get(upload_type, [])
    # asserting that this will return a list, which is not necessarily true
    # check cidc-schemas/schemas/assays/components/available_assays.json

    if isinstance(assay_instances, dict):
        # only exception to list, eg olink
        assay_md = assay_instances
    elif isinstance(assay_instances, list):
        ds = DeepSearch(assay_instances, file_record.object_url)
        if "matched_values" in ds:
            if len(ds["matched_values"]) != 1:
                raise Exception(
                    f"Issue loading antibodies for {file_record.object_url} in {file_record.trial_id}: {file_record.object_url} is not unique in ct['assays'][{upload_type}]"
                )

            # matched_value = ["root[path][to][matching]"]
            matching_path = list(ds["matched_values"])[0]
            index = matching_path.split("[")[1].split("]")[0]
            if index.isdigit():  # not technically needed, see below
                assay_md = assay_instances[int(index)]
            else:
                # technically can't get here because DeepSearch on assay_instances: list has return bounded to "root[ int ]..."
                # if some error occurs, need to error or need assay_md defined
                # testing this doesn't seem necessary, but would likely need patching DeepSearch
                try:
                    assay_md = assay_instances[index]  # should work for all root[...]
                except:
                    # add a bit of actual context, as any IndexError thrown would not be useful
                    raise Exception(
                        f"Issue loading antibodies for {file_record.object_url} in {file_record.trial_id}: unable to search ct['assays']['{upload_type}']"
                    )

    else:
        raise TypeError(
            f"Issue loading antibodies for {file_record.object_url} in {file_record.trial_id}: ct['assays']['{upload_type}'] is {type(assay_instances).__name__} not list, dict"
        )

    md = transforms[upload_type](assay_md)
    if md is None:  # no antibody metadata on the assay
        return None

    file_md = file_record.additional_metadata
    if upload_type == "ihc":
        # for ihc, is only a single antibody
        file_md["ihc.antibody"] = md
    else:
        file_md[f"{upload_type}.antibodies"] = md

    return file_md


def _cytof_antibody_md(assay_md: dict) -> Optional[str]:
    antibody_md = assay_md.get("cytof_antibodies")
    if not antibody_md:
        return None

    antibodies = []
    for ab in antibody_md:
        if ab["usage"] != "Ignored":
            entry = f"{ab['stain_type'].lower().split()[0]} {ab['isotope']}-{ab['antibody']}"
            if ab.get("clone"):
                entry += f" ({ab['clone']})"
            antibodies.append(entry)

    return ", ".join(antibodies)


def _elisa_antibody_md(assay_md: dict) -> Optional[str]:
    antibody_md = assay_md.get("antibodies")
    if not antibody_md:
        return None

    antibodies = []
    for ab in antibody_md:
        if ab["usage"] != "Ignored":
            entry = f"{ab['stain_type'].lower().split()[0]} {ab['isotope']}-{ab['antibody']}"
            if ab.get("clone"):
                entry += f" ({ab['clone']})"
            antibodies.append(entry)

    return ", ".join(antibodies)


def _ihc_antibody_md(assay_md: dict) -> Optional[str]:
    antibody_md = assay_md.get("antibody")
    if not antibody_md:
        return None

    antibody = antibody_md["antibody"]
    if antibody_md.get("clone"):
        antibody += f" ({antibody_md['clone']})"

    return antibody


def _micsss_antibody_md(assay_md: dict) -> Optional[str]:
    antibody_md = assay_md.get("antibody")
    if not antibody_md:
        return None

    antibodies = []
    for ab in antibody_md:
        entry = ab["antibody"]
        if ab.get("clone"):
            entry += f" ({ab['clone']})"
        antibodies.append(entry)

    return ", ".join(antibodies)


def _mif_antibody_md(assay_md: dict) -> Optional[str]:
    antibody_md = assay_md.get("antibodies")
    if not antibody_md:
        return None

    antibodies = []
    for ab in antibody_md:
        if ab.get("export_name"):
            entry = ab["export_name"]
        else:
            entry = ab["antibody"] + " ("
            if ab.get("clone"):
                entry += ab["clone"] + " - "
            entry += str(ab["fluor_wavelength"]) + ")"

        antibodies.append(entry)

    return ", ".join(antibodies)


def _ihc_combined_transform(
    file_record: DownloadableFiles, metadata_df: pd.DataFrame
) -> Optional[dict]:
    """
    Prepare an IHC combined file for visualization by joining it with relevant metadata
    """
    if file_record.upload_type.lower() != "ihc marker combined":
        return None

    print(f"Generating IHC combined visualization config for file {file_record.id}")
    data_file = get_blob_as_stream(file_record.object_url)

    data_df = pd.read_csv(data_file)
    full_df = data_df.join(metadata_df, on="cimac_id", how="inner")

    return json.loads(full_df.to_json(orient="records"))


class _ClustergrammerTransform:
    def __call__(
        self, file_record: DownloadableFiles, metadata_df: pd.DataFrame
    ) -> Optional[dict]:
        """
        Prepare the data file for visualization in clustergrammer. 
        NOTE: `metadata_df` should contain data from the participants and samples CSVs
        for this file's trial, joined on CIMAC ID and indexed on CIMAC ID.
        """
        if file_record.object_url.endswith("npx.xlsx"):
            data_file = get_blob_as_stream(file_record.object_url)
            return self.npx(data_file, metadata_df)
        elif file_record.upload_type.lower() in (
            "cell counts compartment",
            "cell counts assignment",
            "cell counts profiling",
        ):
            data_file = get_blob_as_stream(file_record.object_url)
            return self.cytof_summary(data_file, metadata_df)

        return None

    def npx(self, data_file, metadata_df: pd.DataFrame) -> dict:
        """Prepare an NPX file for visualization in clustergrammer"""
        # Load the NPX data into a dataframe.
        npx_df = _npx_to_dataframe(data_file)

        return self._clustergrammerify(npx_df, metadata_df)

    def cytof_summary(self, data_file, metadata_df: pd.DataFrame) -> dict:
        """Prepare CyTOF summary csv for visualization in clustergrammer"""
        # Load the CyTOF summary data into a dataframe
        cytof_df = _cytof_summary_to_dataframe(data_file)
        return self._clustergrammerify(cytof_df, metadata_df)

    def _clustergrammerify(
        self, data_df: pd.DataFrame, metadata_df: pd.DataFrame
    ) -> dict:
        """
        Produce the clustergrammer config for the given data and metadata dfs.
        `data_df` must be a dataframe with CIMAC ID column headers.
        """
        assert (
            data_df.shape[1] > 1
        ), "Cannot generate clustergrammer visualization for data with only one sample."

        data_df.columns = _metadata_to_categories(metadata_df.loc[data_df.columns])

        # TODO: find a better way to handle missing values
        data_df.fillna(0, inplace=True)

        # Produce a clustergrammer JSON blob for this dataframe.
        net = CGNetwork()
        net.load_df(data_df)
        net.normalize()
        net.cluster()
        return net.viz


def _metadata_to_categories(metadata_df: pd.DataFrame) -> list:
    """
    Add category information to `data_df`'s column headers in the format that Clustergrammer expects:
        "([Category 1]: [Value 1], [Category 2]: [Value 2], ...)"
    """
    metadata_df = metadata_df.copy()  # so don't modify original

    CLINICAL_FIELD_PREFIX = "arbitrary_trial_specific_clinical_annotations."
    columns = []
    for c in metadata_df.columns:
        # go through and check cardinality = # unique
        # also rename the columns to pretty things
        cardinality = len(metadata_df[c].unique())
        if (
            cardinality > CLUSTERGRAMMER_MAX_CATEGORY_CARDINALITY
            or cardinality <= 1
            or cardinality == metadata_df.shape[0]
        ):
            # only want if not all the same, not too many, and not each unique to sample

            if c not in [
                "cimac_participant_id",
                "cohort_name",
                "collection_event_name",
            ]:
                # we want to keep the above no matter what
                metadata_df.pop(c)
                continue

        if "(1=Yes,0=No)" in c:
            # these are boolean! let's treat them that way
            metadata_df[c] = metadata_df[c].astype(bool)

        if c.startswith(CLINICAL_FIELD_PREFIX):
            # for 10021 participants.csv:
            ## remove the prefix
            ## remove any parentheses

            cat = c[len(CLINICAL_FIELD_PREFIX) :]
            if "(" in cat and ")" in cat and cat.index(")") > cat.index("("):
                cat = cat.split("(", 1)[0] + cat.rsplit(")", 1)[1]
        else:
            # otherwise
            ## break up underscores
            ## title case
            ## drop 'CIDC' / 'CIMAC' anywhere
            ## drop trailing 'Name'
            cat = c.replace("_", " ").title().replace("Cidc", "").replace("Cimac", "")
            if cat.endswith("Name") and not cat == "Name":
                cat = cat[:-4]

        # strip so it's pretty!
        if cat.strip() not in columns:
            columns.append(cat.strip())
        else:
            # if it's a repeated name, pop it
            metadata_df.pop(c)

    metadata_df.columns = columns
    print("CG Category options:", ", ".join(columns))

    # cut down to only the categories we want
    columns = [
        c
        for c in [
            "Participant Id",
            "Collection Event",
            "Cohort",
            "Treatment",
            "Disease progression",
            "RECIST clinical benefit status",
        ]
        if c in metadata_df.columns
    ]
    columns = sorted(columns, key=lambda c: len(metadata_df[c].unique()))
    metadata_df = metadata_df[columns]

    if "Disease progression" in columns:
        columns[columns.index("Disease progression")] = "Disease prog"
    if "RECIST clinical benefit status" in columns:
        columns[columns.index("RECIST clinical benefit status")] = "Clin benefit"
    metadata_df.columns = columns

    # build the output str in ClusterGrammer compatible format
    categories = []
    for idx, row in metadata_df.iterrows():
        temp = [f"CIMAC Id: {idx}"]

        for cat, val in row.items():
            temp.append(f"{cat}: {val}")

        categories.append(tuple(temp))

    return categories


def _npx_to_dataframe(fname, sheet_name="NPX Data") -> pd.DataFrame:
    """Load raw data from an NPX file into a pandas dataframe."""

    wb = load_workbook(fname)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Couldn't locate expected worksheet '{sheet_name}'.")
    ws = wb[sheet_name]

    extract_values = lambda xlsx_row: [cell.value for cell in xlsx_row]

    # Assay labels (row 5 of the spreadsheet)
    assay_labels = extract_values(ws[4][1:-2])

    # Raw data (row 8 of the spreadsheet onwards)
    rows = []
    num_cols = len(assay_labels) + 1
    for row in ws.iter_rows(min_row=8):
        sample_id = row[0].value
        # If we hit a blank line, there's no more data to read.
        if not sample_id:
            break
        # Only include rows pertaining to CIMAC ids
        if prism.cimac_id_regex.match(sample_id):
            new_row = extract_values(row[0:num_cols])
            rows.append(new_row)
    raw = pd.DataFrame(rows).set_index(0)
    raw.index.name = "cimac_id"
    raw.columns = assay_labels

    # Drop columns that don't have raw data
    raw.drop(columns=["Plate ID", "QC Warning"], inplace=True)

    # Data is later z-scored, so remove data that would introduce NaN's
    raw.drop(columns=raw.columns[raw.std() == 0], inplace=True)

    return raw.T


def _cytof_summary_to_dataframe(csv: BytesIO) -> pd.DataFrame:
    """Load a CyTOF summary CSV into a dataframe with CIMAC IDs as column headers"""
    raw_df = pd.read_csv(csv)

    # Index on CIMAC ID column
    indexed_df = raw_df.set_index("cimac_id")

    # Drop unused metadata columns (we should get these from the metadata df)
    for col in ["cimac_participant_id", "protocol_identifier"]:
        try:
            indexed_df.drop(col, axis=1, inplace=True)
        except KeyError:
            pass

    return indexed_df.T
